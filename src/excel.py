import os
from openpyxl import *
from datetime import datetime as dt

class Excel():
    def __init__(self,digits):
        try: arr = [dt.now().strftime("%Y/%m/%d"),dt.now().strftime("%H:%M:%S"),"".join(digits)]
        except: arr = [dt.now().strftime("%Y/%m/%d"),dt.now().strftime("%H:%M:%S"),"Error"]
        self.initialize()
        self.saveExcel(arr)

    def initialize(self):
        fPath = os.path.dirname(os.path.dirname(__file__))
        dataPath = os.path.join(fPath,'data')
        self.excelPath = os.path.join(dataPath,dt.today().strftime("%d %b %y")+".xlsx")

    def saveExcel(self,arr):
        if os.path.exists(self.excelPath):
            wb = load_workbook(self.excelPath)
            maxRow = wb.active.max_row
            newRow = maxRow + 1
            for i,j in enumerate(arr): wb.active.cell(row=newRow,column=i+1).value = j
        else:
            wb = Workbook()
            for i,j in enumerate(arr): wb.active.cell(row=1,column=i+1).value = j

        wb.save(self.excelPath)